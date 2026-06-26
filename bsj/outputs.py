"""
Salidas del pipeline: GeoJSON (para mapa / QGIS) y planilla Excel.

GeoJSON usa orden (lon, lat) y CRS WGS84 (EPSG:4326), que es lo que esperan
Leaflet, Google Earth, QGIS, etc. Cada pedimento se exporta como un Polygon
si tiene >=3 vértices, o como Point (centroide) si tiene 1-2.
"""

import json
from datetime import date
from .coords import poligono_a_wgs84, gk_a_wgs84


def pedimentos_a_geojson(pedimentos, datum="posgar2007", fecha_boletin=None):
    feats = []
    for p in pedimentos:
        if not p.vertices:
            continue
        props = {
            "expediente": p.expediente,
            "titular": p.titular,
            "mineral": p.mineral,
            "departamento": p.departamento,
            "superficie_ha": p.superficie_ha,
            "confianza": p.confianza,
            "datum": datum,
            "fecha_boletin": str(fecha_boletin or date.today()),
        }
        if len(p.vertices) >= 3:
            anillo, (clat, clon) = poligono_a_wgs84(p.vertices, datum)
            anillo = anillo + [anillo[0]]  # cerrar el polígono
            geom = {"type": "Polygon", "coordinates": [[[lon, lat] for lon, lat in anillo]]}
            props["centroide_lat"] = round(clat, 6)
            props["centroide_lon"] = round(clon, 6)
        else:
            lat, lon = gk_a_wgs84(*p.vertices[0], datum)
            geom = {"type": "Point", "coordinates": [lon, lat]}
            props["centroide_lat"] = round(lat, 6)
            props["centroide_lon"] = round(lon, 6)
        feats.append({"type": "Feature", "properties": props, "geometry": geom})
    return {"type": "FeatureCollection",
            "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}},
            "features": feats}


def guardar_geojson(pedimentos, ruta, datum="posgar2007", fecha_boletin=None):
    gj = pedimentos_a_geojson(pedimentos, datum, fecha_boletin)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(gj, f, ensure_ascii=False, indent=2)
    return len(gj["features"])


def guardar_xlsx(pedimentos, ruta, datum="posgar2007", fecha_boletin=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedimentos"
    cols = ["Fecha boletín", "Expediente", "Titular", "Mineral", "Departamento",
            "Superficie (ha)", "N° vértices", "Centroide lat", "Centroide lon",
            "Datum", "Confianza"]
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill

    for p in pedimentos:
        clat = clon = None
        if p.vertices:
            if len(p.vertices) >= 3:
                _, (clat, clon) = poligono_a_wgs84(p.vertices, datum)
            else:
                clat, clon = gk_a_wgs84(*p.vertices[0], datum)
        ws.append([
            str(fecha_boletin or date.today()), p.expediente, p.titular, p.mineral,
            p.departamento, p.superficie_ha, len(p.vertices),
            round(clat, 6) if clat is not None else None,
            round(clon, 6) if clon is not None else None,
            datum, p.confianza,
        ])

    anchos = [14, 18, 30, 12, 16, 14, 11, 14, 14, 14, 11]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ws.max_row - 1


# ---------- salidas del MODELO expediente -> eventos -> titulares ----------
def guardar_modelo_json(expedientes, ruta, meta=None):
    """Modelo rico para el visor: expedientes con eventos agrupados y titulares
    historizados. `meta` (dict) lleva rango, fecha de generación y calendario."""
    doc = {"meta": meta or {}, "expedientes": expedientes}
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
    return len(expedientes)


def modelo_a_geojson(expedientes):
    """Un Feature por expediente, usando la geometría ya convertida a WGS84."""
    feats = []
    for e in expedientes:
        pol = e.get("poligono_wgs84")
        cen = e.get("centroide")
        props = {
            "clave": e.get("clave"),
            "expediente": e.get("expediente"),
            "titular": e.get("titular_actual"),
            "mina": e.get("mina"),
            "minerales": ", ".join(e.get("minerales") or []),
            "departamento": e.get("departamento"),
            "estado": e.get("estado"),
            "estado_label": e.get("estado_label"),
            "superficie_ha": e.get("superficie_ha"),
            "n_eventos": len(e.get("eventos") or []),
            "meses": ", ".join(e.get("meses") or []),
            "datum": e.get("datum"),
        }
        if pol and len(pol) >= 3:
            anillo = [[lon, lat] for lon, lat in pol] + [[pol[0][0], pol[0][1]]]
            geom = {"type": "Polygon", "coordinates": [anillo]}
        elif pol and len(pol) >= 1:
            geom = {"type": "Point", "coordinates": [pol[0][0], pol[0][1]]}
        elif cen:
            geom = {"type": "Point", "coordinates": [cen[1], cen[0]]}
        else:
            continue
        if cen:
            props["centroide_lat"] = round(cen[0], 6)
            props["centroide_lon"] = round(cen[1], 6)
        feats.append({"type": "Feature", "properties": props, "geometry": geom})
    return {"type": "FeatureCollection",
            "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}},
            "features": feats}


def guardar_modelo_geojson(expedientes, ruta):
    gj = modelo_a_geojson(expedientes)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(gj, f, ensure_ascii=False, indent=2)
    return len(gj["features"])


def guardar_modelo_xlsx(expedientes, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"
    cols = ["Expediente", "Titular actual", "Mina", "Mineral(es)", "Departamento",
            "Estado (último evento)", "Eventos", "Apariciones totales",
            "Meses con publicación", "Superficie (ha)", "N° vértices",
            "Centroide lat", "Centroide lon", "Datum"]
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill

    for e in expedientes:
        eventos = e.get("eventos") or []
        cen = e.get("centroide")
        ws.append([
            e.get("expediente") or e.get("clave"),
            e.get("titular_actual"),
            e.get("mina"),
            ", ".join(e.get("minerales") or []),
            e.get("departamento"),
            e.get("estado_label"),
            "; ".join(f"{ev['tipo_label']} (x{ev['apariciones']})" for ev in eventos),
            sum(ev["apariciones"] for ev in eventos),
            ", ".join(e.get("meses") or []),
            e.get("superficie_ha"),
            e.get("n_vertices"),
            round(cen[0], 6) if cen else None,
            round(cen[1], 6) if cen else None,
            e.get("datum"),
        ])

    anchos = [18, 28, 14, 18, 16, 22, 34, 12, 22, 14, 11, 13, 13, 12]
    for i, w in enumerate(anchos, 1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ws.max_row - 1


if __name__ == "__main__":
    from .parser import parsear_boletin
    demo = """
    EDICTO El Juzgado Administrativo de Minas hace saber que, MINERA ANDINA DEL SOL S.A.,
    en Expte. N° 414-0123-2026, solicita permiso de exploración para oro en el departamento
    de Iglesia, superficie 3.500 hectáreas. Coordenadas POSGAR 2007:
    Vértice 1: X=6.650.000,00 Y=2.420.000,00  Vértice 2: X=6.650.000,00 Y=2.430.000,00
    Vértice 3: X=6.640.000,00 Y=2.430.000,00  Vértice 4: X=6.640.000,00 Y=2.420.000,00
    """
    peds = parsear_boletin(demo)
    n1 = guardar_geojson(peds, "/home/claude/boletin_minero/salida_demo.geojson")
    n2 = guardar_xlsx(peds, "/home/claude/boletin_minero/salida_demo.xlsx")
    print(f"GeoJSON: {n1} feature(s) | XLSX: {n2} fila(s)")
